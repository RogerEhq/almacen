import datetime
from glob import escape
from datetime import datetime
from django.contrib.auth.decorators import login_required, user_passes_test
from django.forms import DecimalField, models
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_POST, require_http_methods
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.db.models import Sum, Count, Q, ExpressionWrapper, F
from django.utils import timezone
from decimal import Decimal
from django.contrib.auth import logout
from openpyxl import Workbook
from django.db.models import F

from .forms import ProductForm, StockUpdateForm, ClientForm
from .models import Product, Sale, SaleItem, CashDrawerSession, Supplier, Client, SaleReturnItem, SaleReturn
from django.db.models.functions import ExtractYear, ExtractMonth
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
@login_required
def redirect_after_login(request):
    active_session = CashDrawerSession.objects.filter(user=request.user, end_time__isnull=True).first()
    if active_session:
        return redirect('pos_main')
    return redirect('open_session')


@login_required
def open_session_view(request):
    active_session = CashDrawerSession.objects.filter(user=request.user, end_time__isnull=True).first()
    if active_session:
        return redirect('pos_main')

    if request.method == 'POST':
        try:
            starting_balance = Decimal(request.POST.get('starting_balance', '0.00'))
        except:
            starting_balance = Decimal('0.00')

        CashDrawerSession.objects.create(
            user=request.user,
            starting_balance=starting_balance
        )
        return redirect('pos_main')

    return render(request, 'pos/open_session.html')


@login_required
def close_session_view(request):
    active_session = CashDrawerSession.objects.filter(user=request.user, end_time__isnull=True).first()
    if not active_session:
        return redirect('pos_main')

    cash_sales = active_session.sales.filter(payment_method='cash').aggregate(total_sum=Sum('total_amount'))[
                     'total_sum'] or Decimal('0.00')

    card_sales = active_session.sales.filter(payment_method='card').aggregate(total_sum=Sum('total_amount'))[
                     'total_sum'] or Decimal('0.00')
    expected_balance = cash_sales

    context = {
        'session': active_session,
        'cash_sales': cash_sales,
        'card_sales': card_sales,
        'expected_balance': expected_balance,
    }

    if request.method == 'POST':
        try:
            ending_balance = Decimal(request.POST.get('ending_balance', '0.00'))
        except:
            ending_balance = Decimal('0.00')

        notes = request.POST.get('notes', '')

        active_session.end_time = timezone.now()
        active_session.ending_balance = ending_balance
        active_session.notes = notes
        active_session.save()

        logout(request)
        return redirect('login')

    return render(request, 'pos/close_session.html', context)
@login_required
def pos_view(request):
    cart_items = request.session.get('cart', {})
    cart_total = sum(item['subtotal'] for item in cart_items.values())

    active_session = CashDrawerSession.objects.filter(user=request.user, end_time__isnull=True).first()

    context = {
        'cart_items': cart_items.values(),
        'cart_total': cart_total,
        'active_session': active_session
    }
    return render(request, 'pos/pos_main.html', context)


@login_required
@require_POST
def add_product_view(request):
    sku = request.POST.get('sku', '').strip()

    try:
        product = Product.objects.get(sku=sku)
    except Product.DoesNotExist:
        return HttpResponse(
            '<tr style="color: red;"><td colspan="5">Producto con ese código no existe.</td></tr>'
        )

    if product.stock <= 0:
        return HttpResponse(
            '<tr style="color: red;"><td colspan="5">Producto sin stock disponible.</td></tr>'
        )

    cart = request.session.get('cart', {})
    product_id = str(product.id)
    quantity = cart.get(product_id, {}).get('quantity', 0) + 1

    if quantity > product.stock:
        return HttpResponse(
            f'<tr style="color: orange;"><td colspan="5">Stock máximo alcanzado ({product.stock}).</td></tr>'
        )

    cart[product_id] = {
        'id': product.id,
        'sku': product.sku,
        'name': product.name,
        'price': float(product.price),
        'quantity': quantity,
        'subtotal': float(product.price * quantity),
    }

    request.session['cart'] = cart
    request.session.modified = True

    cart_total = sum(item['subtotal'] for item in cart.values())
    context = {
        'item': cart[product_id],
        'cart_total': cart_total
    }

    return render(request, 'pos/cart_row_and_total.html', context)


@login_required
@require_POST
@transaction.atomic
def checkout_view(request):
    cart = request.session.get('cart', {})
    payment_method = request.POST.get('payment_method', 'cash')

    client_id = request.POST.get('client_id')
    client_instance = None
    if client_id:
        try:
            client_instance = Client.objects.get(id=client_id)
        except Client.DoesNotExist:
            pass

    if not cart:
        return HttpResponse('<p style="color:red;">No hay productos en el carrito.</p>')

    active_session = CashDrawerSession.objects.filter(user=request.user, end_time__isnull=True).first()
    if not active_session:
        return HttpResponse('<p style="color:red;">No hay sesión de caja activa.</p>')

    try:
        sale_items_to_create = []
        cart_total = Decimal(0)

        for product_data in cart.values():
            product_id = product_data['id']
            quantity_sold = product_data['quantity']

            product = Product.objects.select_for_update().get(pk=product_id)

            if quantity_sold > product.stock:
                return HttpResponse(f'<p style="color:red;">Stock insuficiente para {escape(product.name)}.</p>')

            product.stock -= quantity_sold
            product.save()

            unit_price = Decimal(product_data['price'])
            subtotal = Decimal(product_data['subtotal'])
            cart_total += subtotal

            sale_items_to_create.append({
                'product': product,
                'quantity': quantity_sold,
                'unit_price': unit_price,
                'subtotal': subtotal
            })

        sale = Sale.objects.create(
            seller=request.user,
            total_amount=cart_total,
            cash_drawer_session=active_session,
            payment_method=payment_method,
            client=client_instance
        )

        SaleItem.objects.bulk_create([
            SaleItem(
                sale=sale,
                product=item['product'],
                quantity=item['quantity'],
                unit_price=item['unit_price'],
                subtotal=item['subtotal']
            ) for item in sale_items_to_create
        ])

        if payment_method == 'cash':
            active_session.starting_balance += cart_total
            active_session.save()

        del request.session['cart']
        request.session.modified = True

    except Exception as e:
        return HttpResponse(f'<p style="color:red;">Error al finalizar la venta: {escape(str(e))}</p>')

    return HttpResponse(f"""
        <div class="success-message">
            ✅ Venta completada con éxito.<br>
            Total vendido: <b>${cart_total:.2f}</b><br>
            Método de pago: <b>{payment_method.title()}</b>
        </div>
        <script>
            const method = "{payment_method}";
            const total = {float(cart_total)};
            const balanceEl = document.getElementById("cash-balance");
            if (method === "cash" && balanceEl) {{
                const current = parseFloat(balanceEl.textContent.replace('$', ''));
                const updated = current + total;
                balanceEl.textContent = "$" + updated.toFixed(2);
            }}
        </script>
    """)

@login_required
def get_cart_total_view(request):
    cart_items = request.session.get('cart', {})
    cart_total = sum(item['subtotal'] for item in cart_items.values())

    context = {'cart_total': cart_total}
    return render(request, 'pos/total_fragment.html', context)


def is_admin_staff(user):
    return user.is_staff or user.is_superuser


@login_required
def home_dispatch_view(request):
    if is_admin_staff(request.user):
        return redirect('dashboard')
    else:
        active_session = CashDrawerSession.objects.filter(
            user=request.user,
            end_time__isnull=True
        ).first()

        if active_session:
            return redirect('pos_main')
        else:
            return redirect('open_session')

def is_admin_staff(user):
    return user.is_staff or user.is_superuser


@user_passes_test(is_admin_staff)
@login_required
def dashboard_view(request):
    today = timezone.now().date()

    today_sales = Sale.objects.filter(sale_date__date=today)

    today_metrics = today_sales.aggregate(
        total_sales=Sum('total_amount'),
        num_transactions=Count('id')
    )

    total_sales = today_metrics.get('total_sales') or Decimal('0.00')
    num_transactions = today_metrics.get('num_transactions') or 0
    ticket_average = (total_sales / num_transactions) if num_transactions else Decimal('0.00')

    active_sessions = CashDrawerSession.objects.filter(end_time__isnull=True).count()

    top_products = SaleItem.objects.values('product__name') \
                       .annotate(total_sold=Sum('quantity')) \
                       .order_by('-total_sold')[:5]

    low_stock_count = Product.objects.filter(
        stock__lte=F('low_stock_threshold')
    ).count()

    context = {
        'ventas_hoy': total_sales,
        'transacciones_hoy': num_transactions,
        'ticket_promedio': ticket_average,
        'sesiones_activas': active_sessions,
        'top_products': top_products,
        'low_stock_count': low_stock_count,
    }

    return render(request, 'pos/dashboard.html', context)


@user_passes_test(is_admin_staff)
@login_required
def sales_report_view(request):
    sales = None
    totals = {}
    date_range_str = ""

    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date_inclusive = datetime.strptime(end_date_str, '%Y-%m-%d').date()

                date_range_str = f"{start_date_str} a {end_date_str}"

                sales = Sale.objects.filter(
                    sale_date__date__gte=start_date,
                    sale_date__date__lte=end_date_inclusive
                ).select_related('cash_drawer_session__user').order_by('-sale_date')

                if 'export_excel' in request.POST:
                    return export_sales_excel(request, sales, date_range_str)

                if 'export_pdf' in request.POST:
                    return export_sales_pdf(request, sales, date_range_str)

                report_totals = sales.aggregate(
                    total_sales=Sum('total_amount'),
                    total_transactions=Count('id'),
                    total_cash_sales=Sum('total_amount', filter=Q(payment_method='cash')),
                    total_card_sales=Sum('total_amount', filter=Q(payment_method='card')),
                )

                totals = {
                    'total_sales': report_totals.get('total_sales') or 0.00,
                    'total_transactions': report_totals.get('total_transactions') or 0,
                    'total_cash': report_totals.get('total_cash_sales') or 0.00,
                    'total_card': report_totals.get('total_card_sales') or 0.00,
                    'start_date': start_date_str,
                    'end_date': end_date_str,
                }

            except ValueError:
                pass

    context = {
        'sales': sales,
        'totals': totals,
    }
    return render(request, 'pos/sales_report.html', context)


@user_passes_test(is_admin_staff)
@login_required
def product_list_view(request):
    filter_by = request.GET.get('filter', 'all')

    products = Product.objects.select_related('category', 'supplier').all()

    if filter_by == 'low_stock':
        products = products.filter(stock__lte=5)

    context = {
        'products': products,
        'current_filter': filter_by,
        'low_stock_count': Product.objects.filter(stock__lte=5).count()
    }

    return render(request, 'pos/product_list.html', context)


@user_passes_test(is_admin_staff)
@login_required
def supplier_inventory_view(request, supplier_id):
    supplier = get_object_or_404(Supplier, pk=supplier_id)

    products = Product.objects.filter(supplier=supplier).select_related('category')

    total_stock_value = Decimal('0.00')

    for product in products:
        product_stock = Decimal(product.stock)
        total_stock_value += product.cost * product_stock

    context = {
        'supplier': supplier,
        'products': products,
        'total_stock_value': total_stock_value
    }

    return render(request, 'pos/supplier_inventory.html', context)

@user_passes_test(is_admin_staff)
@login_required
def monthly_summary_view(request):
    monthly_sales = Sale.objects.annotate(
        year=ExtractYear('sale_date'),
        month=ExtractMonth('sale_date')
    ).values('year', 'month').annotate(
        total_amount=Sum('total_amount'),
        total_transactions=Count('id')
    ).order_by('-year', '-month')

    sales_with_average = []
    for summary in monthly_sales:
        num_transactions = summary['total_transactions']
        total_amount = summary['total_amount']

        if num_transactions and total_amount:
            summary['ticket_promedio'] = total_amount / Decimal(num_transactions)
        else:
            summary['ticket_promedio'] = Decimal('0.00')

        sales_with_average.append(summary)

    context = {
        'monthly_sales': sales_with_average
    }

    return render(request, 'pos/monthly_summary.html', context)


def export_sales_pdf(request, sales_queryset, date_range):

    response = HttpResponse(content_type='application/pdf')
    file_name = f"reporte_ventas_{datetime.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    title_text = f"Reporte de Ventas: {date_range}"
    elements.append(Paragraph(title_text, styles['Heading1']))
    elements.append(Spacer(1, 18))

    data = [
        ['ID Venta', 'Fecha', 'Total ($)', 'Vendedor', 'Comprador']
    ]

    for sale in sales_queryset:
        vendedor_nombre = (
            sale.cash_drawer_session.user.username
            if sale.cash_drawer_session and sale.cash_drawer_session.user
            else 'N/A'
        )

        if sale.client:
            comprador_nombre = str(sale.client)
        else:
            comprador_nombre = "Consumidor Final"

        data.append([
            sale.id,
            sale.sale_date.strftime("%Y-%m-%d %H:%M"),
            f"{sale.total_amount:.2f}",
            vendedor_nombre,
            comprador_nombre,
        ])

    table = Table(data, colWidths=[60, 140, 80, 100, 120])

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4361ee')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ])

    row_count = len(data)
    for i in range(1, row_count):
        bg_color = colors.white if i % 2 == 0 else colors.HexColor('#f0f0f0')
        style.add('BACKGROUND', (0, i), (-1, i), bg_color)

    table.setStyle(style)
    elements.append(table)

    doc.build(elements)

    return response


def export_sales_excel(request, sales_queryset, date_range):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    file_name = f"reporte_ventas_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    headers = ['ID Venta', 'Fecha y Hora', 'Total ($)', 'Método de Pago', 'Vendedor', 'Comprador']
    ws.append(headers)

    header_style = ws['A1':'F1']
    from openpyxl.styles import Font, PatternFill
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="4361ee")

    for cell in header_style[0]:
        cell.font = font
        cell.fill = fill

    for sale in sales_queryset:
        if sale.client:
            comprador_nombre = str(sale.client)
        else:
            comprador_nombre = "Consumidor Final"

        data_row = [
            sale.id,
            sale.sale_date.strftime("%Y-%m-%d %H:%M"),
            sale.total_amount,
            sale.get_payment_method_display(),
            sale.cash_drawer_session.user.username if sale.cash_drawer_session and sale.cash_drawer_session.user else 'N/A',
            comprador_nombre,
        ]
        ws.append(data_row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response


@user_passes_test(is_admin_staff)
@login_required
def low_inventory_alert_view(request):
    low_stock_products = Product.objects.filter(
        stock__lte=F('low_stock_threshold')
    ).order_by('stock')

    context = {
        'products': low_stock_products,
        'alert_count': low_stock_products.count(),
    }

    return render(request, 'pos/low_inventory_alert.html', context)

@user_passes_test(is_admin_staff)
@login_required
def product_edit_view(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        form = StockUpdateForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('low_inventory_alert')
    else:
        form = StockUpdateForm(instance=product)

    return render(request, 'pos/product_form.html', {'form': form, 'product': product})


@login_required
def client_list_view(request):
    clients = Client.objects.all().order_by('last_name')
    context = {'clients': clients}
    return render(request, 'pos/client_list.html', context)



@login_required
def client_create_view(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('client_list')
    else:
        form = ClientForm()

    context = {'form': form, 'title': 'Crear Nuevo Cliente'}
    return render(request, 'pos/client_form.html', context)


@login_required
def client_edit_view(request, client_id):
    client = get_object_or_404(Client, id=client_id)

    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            return redirect('client_list')
    else:
        form = ClientForm(instance=client)

    context = {'form': form, 'client': client, 'title': f'Editar Cliente: {client.first_name}'}
    return render(request, 'pos/client_form.html', context)


@login_required
def client_delete_view(request, client_id):
    client = get_object_or_404(Client, id=client_id)

    if request.method == 'POST':
        client.delete()
        return redirect('client_list')

    context = {'client': client}
    return render(request, 'pos/client_confirm_delete.html', context)


@login_required
def client_search_ajax(request):
    query = request.GET.get('q', '')
    clients_data = []

    if query:
        clients = Client.objects.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(company_name__icontains=query) |
            Q(tax_id__icontains=query)
        ).distinct()[:10]

        for client in clients:
            clients_data.append({
                'id': client.id,
                'text': str(client),
                'tax_id': client.tax_id,
                'is_professional': client.is_professional
            })

    return JsonResponse({'results': clients_data})


@login_required
@require_http_methods(["GET", "POST"])
def return_search_view(request):
    sale = None
    if request.method == 'POST':
        sale_id = request.POST.get('sale_id')
        if sale_id:
            try:
                sale = Sale.objects.get(id=sale_id)
            except Sale.DoesNotExist:
                return render(request, 'pos/return_search.html', {
                    'error': f'No se encontró la venta con ID: {escape(sale_id)}.'
                })

            return redirect('process_return', sale_id=sale.id)

    return render(request, 'pos/return_search.html', {})


@login_required
@require_http_methods(["GET", "POST"])
@transaction.atomic
def process_return_view(request, sale_id):
    sale = get_object_or_404(
        Sale.objects.prefetch_related('items__product'),
        id=sale_id
    )
    if request.method == 'POST':
        items_to_return = []
        total_refund = Decimal('0.00')
        motive = request.POST.get('motive', 'Devolución sin motivo especificado')
        for item in sale.items.all():
            return_qty_str = request.POST.get(f'qty_{item.id}')
            try:
                return_qty = int(return_qty_str) if return_qty_str else 0
            except ValueError:
                return HttpResponse(
                    f'<p style="color:red;">Cantidad inválida para el producto {escape(item.product.name)}.</p>',
                    status=400)
            if return_qty > item.quantity:
                return HttpResponse(
                    f'<p style="color:red;">No se puede devolver más de lo que se compró para {escape(item.product.name)}.</p>',
                    status=400)
            if return_qty > 0:
                refund_amount = item.unit_price * return_qty
                total_refund += refund_amount
                items_to_return.append({
                    'item': item,
                    'quantity': return_qty,
                    'refund_amount': refund_amount
                })
        if not items_to_return:
            return HttpResponse('<p style="color:red;">Debe seleccionar al menos un producto para devolver.</p>',
                                status=400)

        # 1. Crear el registro de devolución (SaleReturn)
        sale_return = SaleReturn.objects.create(
            original_sale=sale,
            returned_by=request.user,
            motive=motive,
            total_refund_amount=total_refund
        )

        # 2. Revertir stock y crear SaleReturnItems
        for data in items_to_return:
            item = data['item']
            SaleReturnItem.objects.create(
                return_request=sale_return,
                product=item.product,
                quantity=data['quantity'],
                refund_amount=data['refund_amount']
            )
            product = Product.objects.get(pk=item.product.pk)
            product.stock += data['quantity']
            product.save()

        # =================================================================
        # MODIFICACIÓN CLAVE: Usar la sesión de caja de la venta original
        # =================================================================

        # Usamos la sesión de la venta original. Si la venta original no tiene sesión (lo cual es raro),
        # usamos la sesión activa actual como fallback, aunque lo ideal es que sale.cash_drawer_session exista.
        session_to_use = sale.cash_drawer_session

        # A. Crear una transacción de VENTA con monto NEGATIVO para compensar las métricas
        Sale.objects.create(
            seller=request.user,
            total_amount=-total_refund,
            # Se usa la sesión de la VENTA ORIGINAL para asegurar que no sea NULL
            cash_drawer_session=session_to_use,
            payment_method='return',
            client=sale.client,
        )

        # B. Actualizar el balance de caja (si el pago original fue en efectivo y tenemos una sesión)
        if sale.payment_method == 'cash' and session_to_use:
            # Si el pago original fue en efectivo, el dinero de la devolución se saca de la caja
            # Si la sesión usada es la original, esta podría estar cerrada.
            # Es más seguro usar la sesión activa del usuario para actualizar el balance de caja si aún está abierta.
            active_session = CashDrawerSession.objects.filter(user=request.user, end_time__isnull=True).first()
            if active_session:
                active_session.starting_balance -= total_refund
                active_session.save()
        # =================================================================

        return render(request, 'pos/return_success.html', {
            'sale_return': sale_return,
            'items_returned': items_to_return
        })
    context = {
        'sale': sale,
        'sale_items': sale.items.all(),
    }
    return render(request, 'pos/process_return.html', context)